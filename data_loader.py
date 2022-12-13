import os
import re
import pickle
import codecs
import openpyxl
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
import pydicom
import nibabel as nib
from torch.utils.data import Dataset
from torchvision import transforms
import torch.nn.functional as F

from utils import log_utils


class ImageDataset(Dataset):
    def __init__(self, pixels, labels, input_size, input_norm=False, do_train=True, debug=False):
        super(ImageDataset, self).__init__()
        self.input_size = input_size

        if do_train:
            self.data_transform = transforms.Compose([
                # transforms.CenterCrop(image_size),
                # transforms.Resize((image_size, image_size)),
                # transforms.RandomCrop(image_size, pad_if_needed=True),
                # transforms.RandomHorizontalFlip(),
                transforms.ToTensor(),
                # transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
            ])
        else:
            self.data_transform = transforms.Compose([
                # transforms.CenterCrop(image_size),
                # transforms.RandomCrop(image_size, pad_if_needed=True),
                # transforms.Resize((image_size, image_size)),
                # transforms.CenterCrop(image_size),
                transforms.ToTensor(),
                # transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
            ])

        data = []
        for id in labels:
            label = labels[id]
            pixel = pixels[id]
            for item in pixel:
                item_max = np.max(item.shape)
                if item_max > input_size:
                    continue
                data.append([item, label])

            if debug and len(data) >= 100:
                break

        if input_norm:
            max_value = 0
            for record in data:
                pixel, label = record
                pixel_max = np.max(pixel)
                if pixel_max >= max_value:
                    max_value = pixel_max

            self.data = []
            for record in data:
                pixel, label = record
                pixel = pixel / max_value
                self.data.append([pixel, label])
        else:
            self.data = data

    def __len__(self):
        return len(self.data)

    def __getitem__(self, idx):
        pixel, label = self.data[idx]
        pixel = pixel.astype(np.float32)
        pixel = self.data_transform(pixel)

        _, height, width = pixel.shape
        pad_left = int((self.input_size - width) / 2)
        pad_right = self.input_size - width - pad_left
        pad_top = int((self.input_size - height) / 2)
        pad_bottom = self.input_size - height - pad_top
        pixel = F.pad(pixel, (pad_left, pad_right, pad_top, pad_bottom))

        return pixel, label


def load_data(data_path):
    pixels = {}
    labels = {}

    with open('%s/data.pickle' % data_path, 'rb') as fp:
        data = pickle.load(fp)

    for person in data:
        labels[person] = data[person]['label']
        pixels[person] = data[person]['pixels']

    return pixels, labels


def rename_files(data_path):
    person_list = os.listdir(data_path)
    for person in person_list:
        log_utils.info('renaming %s' % person)
        file_list = os.listdir('%s/%s' % (data_path, person))
        for file_name in file_list:
            if file_name[-3:] == 'nii':
                os.rename('%s/%s/%s' % (data_path, person, file_name), '%s/%s/label.nii' % (data_path, person))
            elif file_name[-3:] == 'DCM':
                new_name = re.findall(r'\d+\.DCM', file_name)[0]
                os.rename('%s/%s/%s' % (data_path, person, file_name), '%s/%s/%s' % (data_path, person, new_name))


def build_data(data_path, exp_size):
    data = {}

    # read label
    label_book = openpyxl.load_workbook('%s/origin/label.xlsx' % data_path)
    label_sheet = label_book[label_book.sheetnames[0]]
    row_num = label_sheet.max_row
    for row_idx in range(2, row_num + 1):
        person_idx = label_sheet.cell(row=row_idx, column=1).value
        label = label_sheet.cell(row=row_idx, column=2).value
        data[str(person_idx)] = {
            'pixels': [],
            'label': int(label)
        }

    # read pixels
    person_list = os.listdir('%s/origin/images' % data_path)
    for person in person_list:
        log_utils.info('building %s' % person)
        person_path = '%s/origin/images/%s' % (data_path, person)
        file_list = os.listdir(person_path)
        for file_name in file_list:
            if file_name[-3:] == 'nii':
                nii_data = nib.load('%s/%s' % (person_path, file_name))
                width, height, queue = nii_data.dataobj.shape
                for idx in range(queue):
                    queue_data = nii_data.dataobj[:, :, idx]
                    if np.max(queue_data) > 0:
                        effective_pixels = np.where(queue_data == 1)
                        left = max(np.min(effective_pixels[0]) - exp_size, 0)
                        top = max(np.min(effective_pixels[1]) - exp_size, 0)
                        right = min(np.max(effective_pixels[0]) + exp_size, width)
                        bottom = min(np.max(effective_pixels[1]) + exp_size, width)

                        # plt.imshow(queue_data, cmap='gray')
                        # plt.savefig('%s/n_%s.png' % (output_path, idx + 1))
                        # plt.imshow(queue_data[left: right, top: bottom], cmap='gray')
                        # plt.savefig('%s/n_c_%s.png' % (output_path, idx + 1))

                        dcm_data = pydicom.read_file('%s/%s.DCM' % (person_path, idx + 1))
                        dcm_crop = dcm_data.pixel_array[top: bottom, left: right]

                        # plt.imshow(dcm_data.pixel_array, cmap='gray')
                        # plt.savefig('%s/d_%s.png' % (output_path, idx + 1))
                        # plt.imshow(dcm_crop, cmap='gray')
                        # plt.savefig('%s/d_c_%s.png' % (output_path, idx + 1))

                        data[person]['pixels'].append(dcm_crop)

    with open('%s/data.pickle' % data_path, 'wb') as fp:
        pickle.dump(data, fp)

    log_utils.info('complete build effect data')


def check_data_size(data_path):
    size_stat = {30: 0, 50: 0, 100: 0, 150: 0, 180: 0, 190: 0, 200: 0, 300: 0, 400: 0}

    with open('%s/data.pickle' % data_path, 'rb') as fp:
        data = pickle.load(fp)

    for person in data:
        # log_utils.info('reading: %s' % person)
        pixels = data[person]['pixels']
        for pixel in pixels:
            width, height = pixel.shape
            size_max = max(width, height)
            for pixel_size in size_stat:
                if size_max <= pixel_size:
                    size_stat[pixel_size] = size_stat[pixel_size] + 1
                    break

    log_utils.info('pixel size statistic: %s' % size_stat)


if __name__ == '__main__':
    data_path = './data'
    exp_size = 10

    # rename_files('%s/src/images' % data_path)
    # build_data(data_path, exp_size=exp_size)
    check_data_size(data_path)
